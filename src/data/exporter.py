"""
Data exporter for E-Commerce Data Scraper.
Handles exporting scraped data to various formats including CSV, Excel, and JSON.
"""

import pandas as pd
import json
import os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from utils.logger import get_logger


class DataExporter:
    """Handle data export operations in multiple formats."""
    
    def __init__(self):
        self.logger = get_logger(__name__)
        
    def export_data(self, data, filepath, format_type='csv', **kwargs):
        """
        Export data to specified format.
        
        Args:
            data (pd.DataFrame): Data to export
            filepath (str): Output file path
            format_type (str): Export format ('csv', 'excel', 'json')
            **kwargs: Additional format-specific parameters
            
        Returns:
            bool: True if export successful, False otherwise
        """
        if data.empty:
            self.logger.warning("No data to export")
            return False
            
        try:
            # Ensure directory exists
            Path(filepath).parent.mkdir(parents=True, exist_ok=True)
            
            # Export based on format
            if format_type.lower() == 'csv':
                return self._export_csv(data, filepath, **kwargs)
            elif format_type.lower() == 'excel':
                return self._export_excel(data, filepath, **kwargs)
            elif format_type.lower() == 'json':
                return self._export_json(data, filepath, **kwargs)
            else:
                self.logger.error(f"Unsupported export format: {format_type}")
                return False
                
        except Exception as e:
            self.logger.error(f"Export failed: {str(e)}")
            return False
            
    def _export_csv(self, data, filepath, **kwargs):
        """Export data to CSV format."""
        try:
            # Default CSV parameters
            csv_params = {
                'index': False,
                'encoding': 'utf-8-sig',  # BOM for Excel compatibility
                'sep': kwargs.get('delimiter', ','),
                'quoting': 1,  # Quote all non-numeric fields
                'escapechar': '\\'
            }
            
            # Handle header option
            if not kwargs.get('include_headers', True):
                csv_params['header'] = False
                
            # Export data
            data.to_csv(filepath, **csv_params)
            
            self.logger.info(f"Data exported to CSV: {filepath}")
            return True
            
        except Exception as e:
            self.logger.error(f"CSV export failed: {str(e)}")
            return False
            
    def _export_excel(self, data, filepath, **kwargs):
        """Export data to Excel format with formatting."""
        try:
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = kwargs.get('sheet_name', 'Scraped Data')
            
            # Add metadata sheet
            self._add_metadata_sheet(wb, data, kwargs)
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Style header row
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                        
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add filters to header row
            ws.auto_filter.ref = ws.dimensions
            
            # Save workbook
            wb.save(filepath)
            
            self.logger.info(f"Data exported to Excel: {filepath}")
            return True
            
        except Exception as e:
            self.logger.error(f"Excel export failed: {str(e)}")
            return False
            
    def _add_metadata_sheet(self, workbook, data, export_options):
        """Add metadata sheet to Excel workbook."""
        try:
            metadata_ws = workbook.create_sheet("Export Metadata")
            
            # Export information
            metadata = [
                ["Export Information", ""],
                ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Total Records", len(data)],
                ["Columns", len(data.columns)],
                ["Export Format", "Excel"],
                ["", ""],
                ["Column Information", ""],
                ["Column Name", "Data Type"],
            ]
            
            # Add column information
            for col in data.columns:
                metadata.append([col, str(data[col].dtype)])
            
            # Add data to metadata sheet
            for row_idx, row_data in enumerate(metadata, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = metadata_ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Style headers
                    if row_idx == 1 or row_data[0] in ["Export Information", "Column Information"]:
                        cell.font = Font(bold=True)
            
            # Auto-adjust column widths
            for column in metadata_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                        
                adjusted_width = max_length + 2
                metadata_ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.warning(f"Could not create metadata sheet: {str(e)}")
            
    def _export_json(self, data, filepath, **kwargs):
        """Export data to JSON format."""
        try:
            # Prepare data for JSON export
            json_data = {
                'export_info': {
                    'export_date': datetime.now().isoformat(),
                    'total_records': len(data),
                    'columns': list(data.columns)
                },
                'data': data.to_dict('records')
            }
            
            # Handle NaN values
            json_data = self._clean_json_data(json_data)
            
            # Write JSON file
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
            
            self.logger.info(f"Data exported to JSON: {filepath}")
            return True
            
        except Exception as e:
            self.logger.error(f"JSON export failed: {str(e)}")
            return False
            
    def _clean_json_data(self, data):
        """Clean data for JSON serialization by handling NaN values."""
        if isinstance(data, dict):
            return {key: self._clean_json_data(value) for key, value in data.items()}
        elif isinstance(data, list):
            return [self._clean_json_data(item) for item in data]
        elif pd.isna(data):
            return None
        else:
            return data
            
    def export_summary_report(self, data, filepath):
        """
        Export a summary report of the scraped data.
        
        Args:
            data (pd.DataFrame): Scraped data
            filepath (str): Output file path
            
        Returns:
            bool: True if export successful
        """
        try:
            if data.empty:
                self.logger.warning("No data for summary report")
                return False
                
            # Generate summary statistics
            summary = self._generate_summary_stats(data)
            
            # Create summary report
            report_lines = [
                "=" * 60,
                "E-COMMERCE DATA SCRAPER - SUMMARY REPORT",
                "=" * 60,
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                "",
                "DATASET OVERVIEW",
                "-" * 20,
                f"Total Records: {summary['total_records']:,}",
                f"Total Columns: {summary['total_columns']}",
                f"Data Quality Score: {summary['quality_score']:.1f}%",
                "",
                "COLUMN INFORMATION",
                "-" * 20,
            ]
            
            # Add column details
            for col_info in summary['columns']:
                report_lines.append(f"• {col_info['name']}: {col_info['type']} "
                                  f"({col_info['non_null']:,} non-null, "
                                  f"{col_info['null_percentage']:.1f}% missing)")
            
            # Add price analysis if available
            if 'price_stats' in summary:
                report_lines.extend([
                    "",
                    "PRICE ANALYSIS",
                    "-" * 15,
                    f"Average Price: ${summary['price_stats']['mean']:.2f}",
                    f"Price Range: ${summary['price_stats']['min']:.2f} - ${summary['price_stats']['max']:.2f}",
                    f"Median Price: ${summary['price_stats']['median']:.2f}",
                ])
                
            # Add rating analysis if available
            if 'rating_stats' in summary:
                report_lines.extend([
                    "",
                    "RATING ANALYSIS", 
                    "-" * 16,
                    f"Average Rating: {summary['rating_stats']['mean']:.2f}/5.0",
                    f"Rating Range: {summary['rating_stats']['min']:.1f} - {summary['rating_stats']['max']:.1f}",
                    f"Median Rating: {summary['rating_stats']['median']:.2f}",
                ])
            
            # Add category breakdown if available
            if 'category_breakdown' in summary:
                report_lines.extend([
                    "",
                    "CATEGORY BREAKDOWN",
                    "-" * 19,
                ])
                for category, count in summary['category_breakdown'].items():
                    percentage = (count / summary['total_records']) * 100
                    report_lines.append(f"• {category}: {count:,} ({percentage:.1f}%)")
            
            report_lines.append("\n" + "=" * 60)
            
            # Write report to file
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(report_lines))
                
            self.logger.info(f"Summary report exported to: {filepath}")
            return True
            
        except Exception as e:
            self.logger.error(f"Summary report export failed: {str(e)}")
            return False
            
    def _generate_summary_stats(self, data):
        """Generate summary statistics for the dataset."""
        summary = {
            'total_records': len(data),
            'total_columns': len(data.columns),
            'columns': []
        }
        
        # Calculate data quality score
        non_null_percentage = (data.count().sum() / (len(data) * len(data.columns))) * 100
        summary['quality_score'] = non_null_percentage
        
        # Column information
        for col in data.columns:
            col_info = {
                'name': col,
                'type': str(data[col].dtype),
                'non_null': data[col].count(),
                'null_percentage': (data[col].isnull().sum() / len(data)) * 100
            }
            summary['columns'].append(col_info)
            
        # Price statistics
        if 'price' in data.columns:
            price_data = pd.to_numeric(data['price'], errors='coerce').dropna()
            if len(price_data) > 0:
                summary['price_stats'] = {
                    'mean': price_data.mean(),
                    'median': price_data.median(),
                    'min': price_data.min(),
                    'max': price_data.max(),
                    'std': price_data.std()
                }
                
        # Rating statistics
        if 'rating' in data.columns:
            rating_data = pd.to_numeric(data['rating'], errors='coerce').dropna()
            if len(rating_data) > 0:
                summary['rating_stats'] = {
                    'mean': rating_data.mean(),
                    'median': rating_data.median(),
                    'min': rating_data.min(),
                    'max': rating_data.max(),
                    'std': rating_data.std()
                }
                
        # Category breakdown
        if 'category' in data.columns:
            category_counts = data['category'].value_counts()
            summary['category_breakdown'] = category_counts.to_dict()
            
        return summary
        
    def get_supported_formats(self):
        """Get list of supported export formats."""
        return ['csv', 'excel', 'json']
        
    def validate_export_path(self, filepath, format_type):
        """
        Validate export file path and format.
        
        Args:
            filepath (str): Export file path
            format_type (str): Export format
            
        Returns:
            tuple: (is_valid, error_message)
        """
        try:
            # Check if format is supported
            if format_type.lower() not in self.get_supported_formats():
                return False, f"Unsupported format: {format_type}"
                
            # Check if directory is writable
            directory = Path(filepath).parent
            if not directory.exists():
                try:
                    directory.mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    return False, f"Cannot create directory: {str(e)}"
                    
            # Check if file already exists
            if Path(filepath).exists():
                # This is just a warning, not an error
                return True, f"File exists and will be overwritten: {filepath}"
                
            return True, ""
            
        except Exception as e:
            return False, f"Path validation error: {str(e)}"
